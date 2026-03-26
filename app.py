import streamlit as st
import pandas as pd
import numpy as np
import io
import tempfile
import os
from harga_formula_proporsional import (
    proses_penjualan_shopee, _get_bulan_dominan, _rekap_numerik, parse_shopee_filename
)

# ---------------------------------------------------------
# KONFIGURASI HALAMAN
# ---------------------------------------------------------
st.set_page_config(
    page_title="Analisa Harga Shopee",
    page_icon="📊",
    layout="wide"
)

# ---------------------------------------------------------
# CUSTOM CSS
# ---------------------------------------------------------
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    
    /* Global */
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }
    
    /* Header */
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        color: white;
        box-shadow: 0 10px 40px rgba(102, 126, 234, 0.3);
    }
    .main-header h1 {
        margin: 0;
        font-size: 2rem;
        font-weight: 800;
        letter-spacing: -0.5px;
    }
    .main-header p {
        margin: 0.5rem 0 0 0;
        opacity: 0.85;
        font-size: 1rem;
        font-weight: 300;
    }
    
    /* Mode selector cards */
    .mode-card {
        background: linear-gradient(135deg, #f8f9ff 0%, #f0f2ff 100%);
        border: 2px solid #e0e4ff;
        border-radius: 12px;
        padding: 1.5rem;
        text-align: center;
        transition: all 0.3s ease;
    }
    .mode-card:hover {
        border-color: #667eea;
        box-shadow: 0 4px 20px rgba(102, 126, 234, 0.15);
    }
    .mode-card h3 {
        margin: 0.5rem 0 0.3rem 0;
        color: #333;
        font-weight: 700;
    }
    .mode-card p {
        margin: 0;
        color: #666;
        font-size: 0.85rem;
    }
    
    /* Info box */
    .info-box {
        background: linear-gradient(135deg, #e8f4fd 0%, #d1ecf9 100%);
        border-left: 4px solid #667eea;
        border-radius: 0 8px 8px 0;
        padding: 1rem 1.2rem;
        margin: 1rem 0;
        font-size: 0.9rem;
        color: #2c3e50;
    }
    
    /* Result header */
    .result-header {
        background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%);
        border: 1px solid #bbf7d0;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        margin: 1.5rem 0 1rem 0;
    }
    .result-header h3 {
        margin: 0;
        color: #166534;
        font-weight: 700;
    }
    .result-header p {
        margin: 0.3rem 0 0 0;
        color: #15803d;
        font-size: 0.85rem;
    }
    
    /* Table styling */
    .dataframe {
        font-size: 0.85rem !important;
    }
    
    /* Custom HTML Table */
    .custom-table-wrapper {
        max-height: 700px;
        overflow-y: auto;
        border-radius: 8px;
        border: 1px solid rgba(128, 128, 128, 0.2);
        margin: 0.5rem 0 1rem 0;
    }
    .custom-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.85rem;
    }
    .custom-table thead {
        position: sticky;
        top: 0;
        z-index: 1;
    }
    .custom-table th {
        background: linear-gradient(135deg, #4a5568 0%, #2d3748 100%);
        color: #ffffff;
        font-size: 0.9rem;
        font-weight: 700;
        padding: 12px 10px;
        text-align: left;
        white-space: nowrap;
        border-bottom: 3px solid #667eea;
    }
    .custom-table td {
        padding: 8px 10px;
        border-bottom: 1px solid rgba(128, 128, 128, 0.15);
        color: inherit;
    }
    .custom-table tbody tr:hover {
        background-color: rgba(102, 126, 234, 0.08);
    }
    .custom-table tbody tr:nth-child(even) {
        background-color: rgba(128, 128, 128, 0.05);
    }
    .custom-table .cell-unavailable {
        background-color: rgba(220, 38, 38, 0.15);
        color: #f87171;
        font-weight: 600;
    }
    
    /* Upload area */
    [data-testid="stFileUploader"] {
        border-radius: 12px;
    }
    
    /* Metric cards in columns */
    [data-testid="stMetric"] {
        background: linear-gradient(135deg, #f8f9ff 0%, #f0f2ff 100%);
        border: 1px solid #e0e4ff;
        border-radius: 12px;
        padding: 1rem;
    }
    
    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f8f9ff 0%, #eef0ff 100%);
    }
    
    /* Divider */
    .custom-divider {
        height: 2px;
        background: linear-gradient(90deg, transparent, #667eea, transparent);
        margin: 1.5rem 0;
        border: none;
    }
    
    /* Footer */
    .footer {
        text-align: center;
        color: #aaa;
        font-size: 0.8rem;
        margin-top: 3rem;
        padding: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------
# HELPER: Render DataFrame sebagai HTML table custom
# ---------------------------------------------------------
def _render_html_table(df, highlight_col=None):
    """Render DataFrame sebagai HTML table dengan header besar & bold."""
    html = '<div class="custom-table-wrapper"><table class="custom-table">'
    
    # Header
    html += '<thead><tr>'
    for col in df.columns:
        html += f'<th>{col}</th>'
    html += '</tr></thead>'
    
    # Body
    html += '<tbody>'
    for _, row in df.iterrows():
        html += '<tr>'
        for col in df.columns:
            val = str(row[col])
            if highlight_col and col == highlight_col and val == "tidak tersedia":
                html += f'<td class="cell-unavailable">{val}</td>'
            else:
                html += f'<td>{val}</td>'
        html += '</tr>'
    html += '</tbody></table></div>'
    
    st.markdown(html, unsafe_allow_html=True)

# ---------------------------------------------------------
# HELPER: Format Kolom Excel (Openpyxl)
# ---------------------------------------------------------
def _format_excel_sheet(worksheet, df):
    """Memberikan format pada file Excel (koma ribuan & persen) agar bisa di-kalkulasi."""
    from openpyxl.utils import get_column_letter
    
    for col_idx, col_name in enumerate(df.columns, 1):
        # Set lebar kolom
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 22
        
        # Mengecek apakah nama kolom mengandung persen
        is_pct = "%" in str(col_name)
        
        # Iterasi dari baris ke-2 (karena baris 1 adalah header tabel pandas)
        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            # Terapkan format hanya jika isi sel adalah angka
            if isinstance(cell.value, (int, float)):
                if is_pct:
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0'

# ---------------------------------------------------------
# CACHED PROCESSING — Heavy computation hanya sekali per file
# ---------------------------------------------------------
@st.cache_data(show_spinner=False)
def _proses_dan_rekap(file_bytes, filename):
    """
    Proses file pesanan dan kembalikan rekap numerik + bulan dominan.
    Di-cache berdasarkan isi file — tidak akan diproses ulang jika file sama.
    """
    suffix = os.path.splitext(filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    
    try:
        df = proses_penjualan_shopee(tmp_path)
        bulan = _get_bulan_dominan(df)
        rekap = _rekap_numerik(df)
        return rekap, bulan
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

# ---------------------------------------------------------
# FORMAT HELPERS — Operasi ringan, jalan setiap kali UI berubah
# ---------------------------------------------------------
def _format_top_produk(rekap, bulan, urut_berdasarkan, top_persen):
    """Format tabel Top Produk dari data rekap yang sudah di-cache."""
    rekap = rekap.copy()
    
    if urut_berdasarkan == "qty":
        rekap.sort_values(by="Total_Qty_Terjual", ascending=False, inplace=True)
    else:
        rekap.sort_values(by="Total_Omzet_Produk", ascending=False, inplace=True)
    
    top_n = max(1, int(len(rekap) * top_persen / 100))
    output = rekap.head(top_n).copy()
    
    col_harga = f"Rata2 Harga Jual {bulan}"
    output = output[["Nomor Referensi SKU", "Nama Produk", "Total_Omzet_Produk",
                      "Total_Qty_Terjual", "Rata2_Harga_Jual"]].copy()
    
    output.rename(columns={
        "Nomor Referensi SKU": "SKU",
        "Total_Omzet_Produk": "Total Omzet per Produk",
        "Total_Qty_Terjual": "Qty Terjual",
        "Rata2_Harga_Jual": col_harga
    }, inplace=True)
    
    # Simpan versi raw (numerik) untuk di-download ke Excel/CSV
    output_raw = output.copy()
    output_raw["Qty Terjual"] = output_raw["Qty Terjual"].astype(int)
    
    # Format angka untuk tampilan HTML
    output["Qty Terjual"] = output["Qty Terjual"].astype(int)
    output["Total Omzet per Produk"] = output["Total Omzet per Produk"].apply(lambda x: f"{x:,.0f}")
    output[col_harga] = output[col_harga].apply(lambda x: f"{x:,.0f}")
    
    return output, output_raw

def _format_before_after(rekap_before, rekap_after, bulan_before, bulan_after, urut_berdasarkan, top_persen):
    """Format tabel Before-After dari data rekap yang sudah di-cache."""
    rekap_before = rekap_before.copy()
    rekap_after = rekap_after.copy()
    
    # Urutkan & tentukan kolom metrik
    if urut_berdasarkan == "qty":
        rekap_before.sort_values(by="Total_Qty_Terjual", ascending=False, inplace=True)
        col_metrik_src = "Total_Qty_Terjual"
        col_metrik_nama = "Qty Terjual"
    else:
        rekap_before.sort_values(by="Total_Omzet_Produk", ascending=False, inplace=True)
        col_metrik_src = "Total_Omzet_Produk"
        col_metrik_nama = "Total Omzet per Produk"
    
    top_n = max(1, int(len(rekap_before) * top_persen / 100))
    top_before = rekap_before.head(top_n).copy()
    
    # Join harga AFTER
    harga_after = rekap_after[["Nomor Referensi SKU", "Rata2_Harga_Jual"]].copy()
    harga_after.rename(columns={"Rata2_Harga_Jual": "Harga_After"}, inplace=True)
    hasil = top_before.merge(harga_after, on="Nomor Referensi SKU", how="left")
    
    # Hitung % Perubahan
    hasil["Persen_Perubahan"] = np.where(
        hasil["Harga_After"].notna() & (hasil["Rata2_Harga_Jual"] > 0),
        ((hasil["Harga_After"] - hasil["Rata2_Harga_Jual"]) / hasil["Rata2_Harga_Jual"]) * 100,
        np.nan
    )
    
    # Nama kolom dinamis
    col_before = f"Rata2 Harga Jual {bulan_before}"
    col_after = f"Rata2 Harga Jual {bulan_after}"
    
    output = hasil[["Nomor Referensi SKU", col_metrik_src, "Nama Produk",
                     "Rata2_Harga_Jual", "Harga_After", "Persen_Perubahan"]].copy()
    
    output.rename(columns={
        "Nomor Referensi SKU": "SKU",
        col_metrik_src: col_metrik_nama,
        "Rata2_Harga_Jual": col_before,
        "Harga_After": col_after,
        "Persen_Perubahan": "% Perubahan"
    }, inplace=True)
    
    # Simpan versi raw (numerik) untuk di-download ke Excel/CSV
    output_raw = output.copy()
    if urut_berdasarkan == "qty":
        output_raw[col_metrik_nama] = output_raw[col_metrik_nama].astype(int)
    # Persen ubah ke scale 1.0 (misal 5% jadi 0.05) agar di excel bisa di set sbg persen
    output_raw["% Perubahan"] = output_raw["% Perubahan"] / 100
    
    # Format angka untuk tampilan HTML
    if urut_berdasarkan == "qty":
        output[col_metrik_nama] = output[col_metrik_nama].astype(int)
    else:
        output[col_metrik_nama] = output[col_metrik_nama].apply(lambda x: f"{x:,.0f}")
    output[col_before] = output[col_before].apply(lambda x: f"{x:,.0f}")
    
    output[col_after] = output[col_after].apply(
        lambda x: "tidak tersedia" if pd.isna(x) else f"{x:,.0f}"
    )
    
    # Weighted average % perubahan
    mask_valid = hasil["Persen_Perubahan"].notna()
    if mask_valid.any():
        rata2 = (
            (hasil.loc[mask_valid, "Persen_Perubahan"] * hasil.loc[mask_valid, "Total_Omzet_Produk"]).sum()
            / hasil.loc[mask_valid, "Total_Omzet_Produk"].sum()
        )
    else:
        rata2 = np.nan
    
    col_persen = f"% Perubahan ({rata2:.2f}%)" if not pd.isna(rata2) else "% Perubahan"
    output.rename(columns={"% Perubahan": col_persen}, inplace=True)
    output_raw.rename(columns={"% Perubahan": col_persen}, inplace=True)
    
    output[col_persen] = output[col_persen].apply(
        lambda x: "-" if pd.isna(x) else f"{x:.2f}%"
    )
    
    return output, output_raw

# ---------------------------------------------------------
# HEADER
# ---------------------------------------------------------
st.markdown("""
<div class="main-header">
    <h1>📊 Analisa Harga Shopee</h1>
    <p>Tool otomatis untuk menganalisa top produk & membandingkan perubahan harga antar bulan</p>
</div>
""", unsafe_allow_html=True)

# ---------------------------------------------------------
# MODE SELECTION
# ---------------------------------------------------------
st.markdown("### 🎯 Pilih Mode Analisa")

mode = st.radio(
    "Mode",
    ["📋 Top Produk (1 File)", "🔄 Bandingkan Harga Before vs After (2 File)"],
    horizontal=True,
    label_visibility="collapsed",
    index=1
)

st.markdown('<div class="custom-divider"></div>', unsafe_allow_html=True)

# =========================================================
# MODE 1: TOP PRODUK
# =========================================================
if mode == "📋 Top Produk (1 File)":
    st.markdown("### 📋 Top Produk — Analisa 1 File")
    st.markdown("""
    <div class="info-box">
        📁 Upload file pesanan Shopee (<b>.xlsx</b>, <b>.csv</b>, atau <b>.zip</b>) untuk melihat produk terlaris.
    </div>
    """, unsafe_allow_html=True)
    
    # Upload DULU, baru settings
    file = st.file_uploader(
        "Upload file pesanan",
        type=["xlsx", "csv", "zip"],
        key="top_produk_upload"
    )
    
    if file:
        # Proses file (CACHED — hanya sekali per file)
        try:
            with st.spinner("⏳ Sedang memproses data..."):
                rekap, bulan = _proses_dan_rekap(file.getvalue(), file.name)
        except Exception as e:
            st.error(f"❌ Error saat memproses file: {str(e)}")
            st.stop()
        
        # Settings SETELAH file diproses (mengubah ini tidak trigger reprocessing)
        col_set1, col_set2 = st.columns(2)
        with col_set1:
            urut = st.selectbox(
                "🔢 Urutkan berdasarkan",
                ["Omzet (Tertinggi)", "Qty Terjual (Terbanyak)"],
                index=0
            )
        with col_set2:
            persen = st.slider("📊 Top produk (%)", min_value=5, max_value=100, value=20, step=5)
        
        # Format tabel (operasi ringan, jalan setiap UI berubah)
        urut_param = "qty" if "Qty" in urut else "omzet"
        hasil_html, hasil_raw = _format_top_produk(rekap, bulan, urut_param, persen)
        # Parse filename
        kode_brand, display_name = parse_shopee_filename(file.name)
        
        # Result header
        st.markdown(f"""
        <div class="result-header">
            <h3>✅ Berhasil! {kode_brand} Top {persen}% Produk ({len(hasil_html)} SKU) — Diurutkan berdasarkan {urut}</h3>
            <p>📄 File: {display_name}</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Display table
        _render_html_table(hasil_html)
        # Download buttons
        st.markdown("<br>", unsafe_allow_html=True)
        col_dl1, col_dl2, _ = st.columns([1, 1, 2])
        
        with col_dl1:
            csv_data = hasil_raw.to_csv(index=False).encode('utf-8')
            st.download_button(
                "📥 Download CSV",
                csv_data,
                file_name=f"top_produk_{file.name.rsplit('.', 1)[0]}.csv",
                mime="text/csv"
            )
            
        with col_dl2:
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                hasil_raw.to_excel(writer, index=False, sheet_name='Top Produk')
                _format_excel_sheet(writer.sheets['Top Produk'], hasil_raw)
            
            st.download_button(
                "📥 Download Excel (XLSX)",
                buffer.getvalue(),
                file_name=f"top_produk_{file.name.rsplit('.', 1)[0]}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# =========================================================
# MODE 2: BANDINGKAN HARGA BEFORE VS AFTER
# =========================================================
else:
    st.markdown("### 🔄 Bandingkan Harga — Before vs After")
    st.markdown("""
    <div class="info-box">
        📁 Upload <b>2 file</b> dari toko yang sama dengan bulan berbeda.<br>
        File <b>BEFORE</b> = bulan acuan (top produk).<br>
        File <b>AFTER</b> = bulan pembanding (hanya dilihat harganya).
    </div>
    """, unsafe_allow_html=True)
    
    # Upload 2 files
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 📅 File BEFORE (Bulan Acuan)")
        file_before = st.file_uploader(
            "Upload file BEFORE",
            type=["xlsx", "csv", "zip"],
            key="before_upload"
        )
    
    with col2:
        st.markdown("#### 📅 File AFTER (Bulan Pembanding)")
        file_after = st.file_uploader(
            "Upload file AFTER",
            type=["xlsx", "csv", "zip"],
            key="after_upload"
        )
    
    if file_before and file_after:
        # Proses kedua file (CACHED — hanya sekali per file)
        try:
            with st.spinner("⏳ Sedang memproses data..."):
                rekap_before, bulan_before = _proses_dan_rekap(file_before.getvalue(), file_before.name)
                rekap_after, bulan_after = _proses_dan_rekap(file_after.getvalue(), file_after.name)
        except Exception as e:
            st.error(f"❌ Error saat memproses file: {str(e)}")
            st.stop()
        
        # Settings SETELAH file diproses (mengubah ini tidak trigger reprocessing)
        col_set1, col_set2 = st.columns(2)
        with col_set1:
            urut_ba = st.selectbox(
                "🔢 Urutkan berdasarkan",
                ["Omzet (Tertinggi)", "Qty Terjual (Terbanyak)"],
                index=0,
                key="ba_urut"
            )
        with col_set2:
            persen = st.slider("📊 Top produk (%)", min_value=5, max_value=100, value=20, step=5, key="ba_topn")
        
        # Format tabel (operasi ringan, jalan setiap UI berubah)
        urut_param_ba = "qty" if "Qty" in urut_ba else "omzet"
        hasil_html, hasil_raw = _format_before_after(rekap_before, rekap_after, bulan_before, bulan_after, urut_param_ba, persen)
        
        # Parse filenames
        kode_brand, display_before = parse_shopee_filename(file_before.name)
        _, display_after = parse_shopee_filename(file_after.name)
        
        # Result header
        st.markdown(f"""
        <div class="result-header">
            <h3>✅ Berhasil! Perbandingan Harga — {kode_brand} Top {persen}% Produk ({len(hasil_html)} SKU)</h3>
            <p>📄 BEFORE: {display_before} → AFTER: {display_after}</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Display table
        col_after_name = [c for c in hasil_html.columns if c.startswith("Rata2 Harga Jual")][1]
        _render_html_table(hasil_html, highlight_col=col_after_name)
        # Download buttons
        st.markdown("<br>", unsafe_allow_html=True)
        col_dl1, col_dl2, _ = st.columns([1, 1, 2])
        
        with col_dl1:
            csv_data = hasil_raw.to_csv(index=False).encode('utf-8')
            st.download_button(
                "📥 Download CSV",
                csv_data,
                file_name=f"perbandingan_{file_before.name.rsplit('.', 1)[0]}_vs_{file_after.name.rsplit('.', 1)[0]}.csv",
                mime="text/csv"
            )
            
        with col_dl2:
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                hasil_raw.to_excel(writer, index=False, sheet_name='Perbandingan Harga')
                _format_excel_sheet(writer.sheets['Perbandingan Harga'], hasil_raw)
                
            st.download_button(
                "📥 Download Excel (XLSX)",
                buffer.getvalue(),
                file_name=f"perbandingan_{file_before.name.rsplit('.', 1)[0]}_vs_{file_after.name.rsplit('.', 1)[0]}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# ---------------------------------------------------------
# FOOTER
# ---------------------------------------------------------
st.markdown('<div class="custom-divider"></div>', unsafe_allow_html=True)
st.markdown("""
<div class="footer">
    🛠️ Dibuat dengan Streamlit & Python | Analisa Harga Shopee Tool
</div>
""", unsafe_allow_html=True)
